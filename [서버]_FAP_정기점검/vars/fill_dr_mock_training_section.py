from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_WORKBOOK = Path('/fap/ansible/projects/[서버]_FAP_정기점검/templates/dr_mock_training_check_2026.xlsx')
TARGET_WORKBOOK = Path('/fap/ansible/projects/[서버]_FAP_정기점검/templates/server_report_form.xlsx')
SOURCE_SHEET_CANDIDATES = ['2026', '통합']
TARGET_SHEET = 'Sheet1'
SOURCE_START_ROW = 4
TARGET_START_ROW = 203
COLUMN_MAP = {
    2: 1,
    3: 2,
    4: 4,
    5: 6,
    7: 7,
}


def choose_source_sheet(workbook):
    for name in SOURCE_SHEET_CANDIDATES:
        if name in workbook.sheetnames:
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def has_meaningful_data(ws, row):
    return any(ws.cell(row, col).value not in (None, '') for col in COLUMN_MAP)


def normalize_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return value


def read_source_rows(ws):
    rows = []
    for row in range(SOURCE_START_ROW, ws.max_row + 1):
        if not has_meaningful_data(ws, row):
            continue
        row_data = {}
        for src_col, dst_col in COLUMN_MAP.items():
            row_data[dst_col] = normalize_date(ws.cell(row, src_col).value)
        rows.append(row_data)
    return rows


def copy_cell_style(ws, src_row, dst_row, col):
    src = ws.cell(src_row, col)
    dst = ws.cell(dst_row, col)
    if src.has_style:
        dst._style = copy(src._style)
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def prepare_target_rows(ws, row_count):
    template_row = TARGET_START_ROW
    if row_count <= 1:
        return
    for _ in range(row_count - 1):
        ws.insert_rows(template_row + 1)
        for col in range(1, ws.max_column + 1):
            copy_cell_style(ws, template_row, template_row + 1, col)
        ws.row_dimensions[template_row + 1].height = ws.row_dimensions[template_row].height
        ws.merge_cells(start_row=template_row + 1, start_column=2, end_row=template_row + 1, end_column=3)
        ws.merge_cells(start_row=template_row + 1, start_column=4, end_row=template_row + 1, end_column=5)
        template_row += 1


def write_target_rows(ws, rows):
    for offset, row_data in enumerate(rows):
        row = TARGET_START_ROW + offset
        ws.cell(row, 1, row_data.get(1))
        ws.cell(row, 2, row_data.get(2))
        ws.cell(row, 4, row_data.get(4))
        ws.cell(row, 6, row_data.get(6))
        ws.cell(row, 7, row_data.get(7))


def fill_dr_mock_training_section(output_path):
    source_wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    source_ws = choose_source_sheet(source_wb)
    source_rows = read_source_rows(source_ws)
    if not source_rows:
        raise RuntimeError('No DR mock training rows were found in the source workbook.')

    target_wb = load_workbook(TARGET_WORKBOOK)
    target_ws = target_wb[TARGET_SHEET]
    prepare_target_rows(target_ws, len(source_rows))
    write_target_rows(target_ws, source_rows)

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(output_file)
    return output_file


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        raise SystemExit('Usage: fill_dr_mock_training_section.py <output_xlsx_path>')
    saved = fill_dr_mock_training_section(sys.argv[1])
    print(saved)
