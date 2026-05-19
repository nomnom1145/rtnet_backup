import json
import re
import sys
import unicodedata
from copy import copy
from datetime import datetime
from math import ceil
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CPU_THRESHOLD = 80.0
DISK_THRESHOLD = 80.0
MEMORY_THRESHOLD = 80.0
MISSING_TEXT = "미수집"
CHECK_NEEDED_TEXT = "확인필요"
NORMAL_TEXT = "정상"
VULNERABLE_TEXT = "취약"
LINE_HEIGHT = 15
LINE_PADDING = 6


def clean_lines(value):
    if not value:
        return []
    if isinstance(value, list):
        lines = value
    else:
        lines = str(value).splitlines()
    return [line.strip() for line in lines if str(line).strip()]


def set_cell_alignment(ws, cell_ref, horizontal="center"):
    cell = ws[cell_ref]
    alignment = copy(cell.alignment)
    alignment.horizontal = horizontal
    alignment.vertical = "center"
    alignment.wrap_text = True
    cell.alignment = alignment
    return cell


def find_merged_range(ws, cell_ref):
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            return merged_range
    return None


def visual_length(text):
    total = 0
    for char in text:
        total += 2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1
    return total


def estimate_wrapped_lines(text, capacity):
    if not text:
        return 1
    capacity = max(1, int(capacity))
    wrapped = 0
    for raw_line in str(text).splitlines() or [""]:
        line = raw_line if raw_line else " "
        wrapped += max(1, ceil(visual_length(line) / capacity))
    return max(1, wrapped)


def auto_adjust_row_height(ws, cell_ref, value):
    text = "" if value is None else str(value)
    if not text:
        return

    merged_range = find_merged_range(ws, cell_ref)
    if merged_range:
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row
    else:
        cell = ws[cell_ref]
        min_col = max_col = cell.column
        min_row = max_row = cell.row

    width_units = 0
    for col_idx in range(min_col, max_col + 1):
        letter = get_column_letter(col_idx)
        width_units += ws.column_dimensions[letter].width or 10
    capacity = max(8, int(width_units * 1.1))
    required_lines = estimate_wrapped_lines(text, capacity)
    required_total_height = (required_lines * LINE_HEIGHT) + LINE_PADDING

    row_indices = list(range(min_row, max_row + 1))
    default_height = ws.sheet_format.defaultRowHeight or 15
    current_heights = [ws.row_dimensions[idx].height or default_height for idx in row_indices]
    current_total_height = sum(current_heights)
    if required_total_height <= current_total_height:
        return

    extra_height = required_total_height - current_total_height
    extra_per_row = extra_height / len(row_indices)
    for idx, current_height in zip(row_indices, current_heights):
        ws.row_dimensions[idx].height = current_height + extra_per_row


def set_result_cell(ws, cell_ref, value):
    cell = set_cell_alignment(ws, cell_ref, "center")
    cell.value = value
    auto_adjust_row_height(ws, cell_ref, value)


def set_left_result_cell(ws, cell_ref, value):
    cell = set_cell_alignment(ws, cell_ref, "left")
    cell.value = value
    auto_adjust_row_height(ws, cell_ref, value)


def find_first(lines, prefix):
    for line in lines:
        if line.startswith(prefix):
            return line.split("=", 1)[1].strip()
    return ""


def extract_percent(value):
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", value or "")
    return float(match.group(1)) if match else None


def judge_threshold(value, threshold):
    if value is None:
        return CHECK_NEEDED_TEXT
    if value >= threshold:
        return VULNERABLE_TEXT
    return NORMAL_TEXT


def parse_cpu_usage(text):
    percent = None
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s+us", text or "")
    if match:
        percent = float(match.group(1))
        display = f"{percent:.1f}".rstrip("0").rstrip(".")
        return f"{display}%", percent
    generic = extract_percent(text)
    if generic is not None:
        return f"{generic:g}%", generic
    return MISSING_TEXT if not text else text, None


def parse_size_to_mb(raw_value):
    if not raw_value:
        return None
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*([KMGT])B?$", str(raw_value).strip(), re.IGNORECASE)
    if not match:
        return None
    size = float(match.group(1))
    unit = match.group(2).upper()
    multipliers = {
        "K": 1.0 / 1024.0,
        "M": 1.0,
        "G": 1024.0,
        "T": 1024.0 * 1024.0,
    }
    return size * multipliers[unit]


def format_gb(raw_value):
    size_mb = parse_size_to_mb(raw_value)
    if size_mb is None:
        return raw_value or MISSING_TEXT
    size_gb = size_mb / 1024.0
    if abs(size_gb - round(size_gb)) < 0.05:
        return f"{int(round(size_gb))}GB"
    return f"{size_gb:.1f}GB"


def parse_disk_rows(lines):
    rows = {}
    for line in lines:
        mount = re.search(r"Mount=([^|]+)", line)
        size = re.search(r"Size=([^|]+)", line)
        used = re.search(r"Used=([^|]+)", line)
        use_pct = re.search(r"Use=([^|]+)", line)
        if not mount:
            continue
        mount_name = mount.group(1).strip()
        rows[mount_name] = {
            "mount": mount_name,
            "size": size.group(1).strip() if size else "",
            "used": used.group(1).strip() if used else "",
            "use_pct": use_pct.group(1).strip() if use_pct else "",
            "use_pct_value": extract_percent(use_pct.group(1).strip()) if use_pct else None,
        }
    return rows


def parse_memory(text):
    total = re.search(r"Total=([^|]+)", text or "")
    used = re.search(r"Used=([^|]+)", text or "")
    total_raw = total.group(1).strip() if total else ""
    used_raw = used.group(1).strip() if used else ""
    total_mb = parse_size_to_mb(total_raw)
    used_mb = parse_size_to_mb(used_raw)
    usage_pct = None
    if total_mb and used_mb is not None and total_mb > 0:
        usage_pct = (used_mb / total_mb) * 100.0
    return {
        "total_display": format_gb(total_raw) if total_raw else MISSING_TEXT,
        "used_display": format_gb(used_raw) if used_raw else MISSING_TEXT,
        "usage_pct": usage_pct,
    }


def parse_interface(lines):
    dns = []
    for line in lines:
        if line.startswith("DNS"):
            _, value = line.split("=", 1)
            dns.append(value.strip())
    return {
        "ip": find_first(lines, "IPADDR=") or MISSING_TEXT,
        "dns": "\n".join(dns) if dns else MISSING_TEXT,
        "if": find_first(lines, "IF=") or MISSING_TEXT,
    }


def normalize_process_status(status_text):
    status_text = (status_text or "").strip()
    if not status_text:
        return MISSING_TEXT
    lowered = status_text.lower()
    if lowered.startswith("up"):
        return "Up"
    if lowered.startswith("exited"):
        return "Exited"
    if lowered.startswith("restarting"):
        return "Restarting"
    if lowered.startswith("created"):
        return "Created"
    if lowered.startswith("paused"):
        return "Paused"
    return status_text.split()[0]


def parse_process(lines):
    result = {}
    for line in lines:
        if line.startswith("NAME") or line.startswith("STATUS"):
            continue
        parts = line.split(None, 1)
        if not parts:
            continue
        name = parts[0]
        status = parts[1] if len(parts) > 1 else ""
        result[name] = normalize_process_status(status)
    return result


def write_disk_section(ws, disk_rows, row_map):
    for mount, row in row_map.items():
        row_data = disk_rows.get(mount)
        if row_data:
            set_result_cell(ws, f"B{row}", row_data["size"] or MISSING_TEXT)
            set_result_cell(ws, f"C{row}", row_data["used"] or MISSING_TEXT)
            set_result_cell(ws, f"D{row}", row_data["use_pct"] or MISSING_TEXT)
            set_result_cell(ws, f"E{row}", judge_threshold(row_data["use_pct_value"], DISK_THRESHOLD))
        else:
            set_result_cell(ws, f"B{row}", MISSING_TEXT)
            set_result_cell(ws, f"C{row}", MISSING_TEXT)
            set_result_cell(ws, f"D{row}", MISSING_TEXT)
            set_result_cell(ws, f"E{row}", CHECK_NEEDED_TEXT)


def write_process_row(ws, cell_map, process_map):
    if process_map:
        for name, cell_ref in cell_map.items():
            set_result_cell(ws, cell_ref, process_map.get(name, MISSING_TEXT))
    else:
        for index, cell_ref in enumerate(cell_map.values()):
            set_result_cell(ws, cell_ref, MISSING_TEXT if index == 0 else "")


def write_host_section(ws, host_data, section):
    cpu_display, cpu_pct = parse_cpu_usage(host_data.get("cpu_output", ""))
    set_result_cell(ws, section["cpu_value"], cpu_display)
    set_result_cell(ws, section["cpu_result"], judge_threshold(cpu_pct, CPU_THRESHOLD))

    disk_rows = parse_disk_rows(clean_lines(host_data.get("disk_output")))
    write_disk_section(ws, disk_rows, section["disk_rows"])

    memory = parse_memory(host_data.get("memory_output", ""))
    set_result_cell(ws, section["memory_total"], memory["total_display"])
    set_result_cell(ws, section["memory_used"], memory["used_display"])
    set_result_cell(ws, section["memory_result"], judge_threshold(memory["usage_pct"], MEMORY_THRESHOLD))

    routing_lines = clean_lines(host_data.get("routing_output"))
    set_left_result_cell(ws, section["routing"], "\n".join(routing_lines) if routing_lines else MISSING_TEXT)

    interface = parse_interface(clean_lines(host_data.get("interface_output")))
    set_result_cell(ws, section["interface_ip"], interface["ip"])
    set_result_cell(ws, section["interface_dns"], interface["dns"])
    set_result_cell(ws, section["interface_if"], interface["if"])

    login_lines = clean_lines(host_data.get("login_output"))
    set_left_result_cell(ws, section["login"], "\n".join(login_lines) if login_lines else MISSING_TEXT)

    process_map = parse_process(clean_lines(host_data.get("process_output")))
    write_process_row(ws, section["process_cells"], process_map)


def main():
    if len(sys.argv) != 4:
        raise SystemExit("usage: generate_network_report.py <template> <input_json> <output_xlsx>")

    template_path = Path(sys.argv[1])
    input_json_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    data = json.loads(input_json_path.read_text(encoding="utf-8-sig"))

    wb = load_workbook(template_path)
    ws = wb["Sheet1"]

    report_date = data.get("report_date") or datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y년 %m월 %d일")
    set_result_cell(ws, "C3", report_date)
    set_result_cell(ws, "F4", data.get("first_name", ""))
    set_result_cell(ws, "F5", data.get("second_name", ""))
    set_left_result_cell(ws, "B15", data.get("special_note", ""))

    upper_section = {
        "cpu_value": "A48",
        "cpu_result": "E48",
        "disk_rows": {"/": 60, "/home": 61, "/data": 62},
        "memory_total": "A76",
        "memory_used": "B76",
        "memory_result": "D76",
        "routing": "A87",
        "interface_ip": "B112",
        "interface_dns": "B113",
        "interface_if": "B114",
        "login": "A127",
        "process_cells": {
            "fap-agent": "B153",
            "fap-manager-0": "C153",
            "fap-web": "D153",
            "pg-0": "E153",
            "fap-haproxy": "F153",
            "redis-0": "G153",
        },
    }

    lower_section = {
        "cpu_value": "A52",
        "cpu_result": "E52",
        "disk_rows": {"/": 66, "/home": 67, "/data": 68},
        "memory_total": "A80",
        "memory_used": "B80",
        "memory_result": "D80",
        "routing": "A97",
        "interface_ip": "B117",
        "interface_dns": "B118",
        "interface_if": "B119",
        "login": "A136",
        "process_cells": {
            "fap-agent": "B156",
            "fap-manager-0": "C156",
            "fap-web": "D156",
            "pg-0": "E156",
            "fap-haproxy": "F156",
            "redis-0": "G156",
        },
    }

    write_host_section(ws, data.get("first_host", {}), upper_section)
    write_host_section(ws, data.get("second_host", {}), lower_section)

    set_result_cell(ws, "F169", f"현재 등록된 호스트 수 : {data.get('host_count', '')}")
    set_result_cell(ws, "F171", f"현재 등록된 사용자 수 : {data.get('user_count', '')}")
    set_result_cell(ws, "F172", f"현재 등록된 스케줄 수 : {data.get('schedule_count', '')}")
    set_result_cell(ws, "F173", f"현재 등록된 통보 수 : {data.get('notification_count', '')}")
    set_result_cell(ws, "F174", f"현재 등록된 작업 템플릿 수 : {data.get('template_count', '')}")

    sftp_lines = clean_lines(data.get("sftp_output"))
    set_result_cell(ws, "F179", "\n".join(sftp_lines) if sftp_lines else MISSING_TEXT)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(str(output_path))


if __name__ == "__main__":
    main()
