import json
import math
import os
import re
import sys
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment

TARGET_SHEET = "Sheet1"
DR_SOURCE_SHEET_CANDIDATES = ["2026", "통합"]
DR_SOURCE_START_ROW = 4
DR_SECTION_TITLE = "10. DR 모의훈련 작업 내역"


def load_input(arg):
    if os.path.isfile(arg):
        with open(arg, "r", encoding="utf-8") as file:
            return json.load(file)
    return json.loads(arg)


def choose_source_sheet(workbook):
    for name in DR_SOURCE_SHEET_CANDIDATES:
        if name in workbook.sheetnames:
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def normalize_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value


def get_writable_cell(ws, row, col):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return cell

    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col)

    return cell


def copy_cell_style(ws, src_row, dst_row, col):
    src = ws.cell(src_row, col)
    dst = ws.cell(dst_row, col)

    if isinstance(dst, MergedCell):
        return

    if src.has_style:
        dst._style = copy(src._style)
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def copy_row_style(ws, src_row, dst_row, max_col=7):
    for col in range(1, max_col + 1):
        copy_cell_style(ws, src_row, dst_row, col)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def clone_row_format(ws, src_row, dst_row, max_col=7, merge_ranges=None):
    """
    Copy the full visual layout from one row to another.

    The row has to be normalized before and after merging so that styles
    survive insertions into ranges that later become merged cells.
    """
    normalize_row_merges(ws, dst_row, merge_ranges=None)
    copy_row_style(ws, src_row, dst_row, max_col=max_col)
    normalize_row_merges(ws, dst_row, merge_ranges=merge_ranges)
    copy_row_style(ws, src_row, dst_row, max_col=max_col)


def ensure_merge(ws, start_row, start_col, end_row, end_col):
    target = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"
    for merged in list(ws.merged_cells.ranges):
        if str(merged) == target:
            return
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def unmerge_row_single_line_ranges(ws, row):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row and merged.min_row == merged.max_row:
            try:
                ws.unmerge_cells(str(merged))
            except (KeyError, ValueError):
                pass


def unmerge_overlapping_ranges(ws, start_row, end_row):
    """
    Remove pre-existing merged blocks that overlap a dynamic table region.

    Some template areas reserve space as one large merged block. If we later
    reuse that band as row-by-row table data, those merges must be cleared
    first or borders/alignment appear to collapse into a single cell block.
    """
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row < start_row or merged.min_row >= end_row:
            continue
        try:
            ws.unmerge_cells(str(merged))
        except (KeyError, ValueError):
            pass


def normalize_row_merges(ws, row, merge_ranges=None):
    unmerge_row_single_line_ranges(ws, row)

    for start_col, end_col in merge_ranges or []:
        ensure_merge(ws, row, start_col, row, end_col)


def clear_row_values(ws, row, max_col=7):
    cleared = set()
    for col in range(1, max_col + 1):
        cell = get_writable_cell(ws, row, col)
        key = (cell.row, cell.column)
        if key in cleared:
            continue
        if not isinstance(cell, MergedCell):
            cell.value = None
            cleared.add(key)


def rebuild_dynamic_rows(ws, start_row, next_section_row, row_count, template_row=None, max_col=7, merge_ranges=None):
    row_count = max(int(row_count or 1), 1)
    template_row = template_row or start_row
    available_rows = max(next_section_row - start_row, 1)
    extra_rows = max(row_count - available_rows, 0)

    unmerge_overlapping_ranges(ws, start_row, next_section_row)

    if extra_rows > 0:
        insert_at = start_row + available_rows
        ws.insert_rows(insert_at, extra_rows)
        for offset in range(extra_rows):
            row = insert_at + offset
            clone_row_format(ws, template_row, row, max_col=max_col, merge_ranges=merge_ranges)

    total_rows = max(available_rows, row_count)
    for offset in range(total_rows):
        row = start_row + offset
        if offset < row_count:
            clone_row_format(ws, template_row, row, max_col=max_col, merge_ranges=merge_ranges)
        clear_row_values(ws, row, max_col=max_col)

    return start_row + total_rows


def find_row_by_first_column(ws, label):
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if str(value or "").strip() == label:
            return row
    raise ValueError(f"Could not find row label: {label}")


def find_row_contains(ws, text, start_row=1, end_row=None, col=1):
    end_row = end_row or ws.max_row
    for row in range(start_row, end_row + 1):
        value = str(ws.cell(row, col).value or "").strip()
        if text in value:
            return row
    raise ValueError(f"Could not find row containing: {text}")


def find_section_row(ws, label):
    for row in range(1, ws.max_row + 1):
        value = str(ws.cell(row, 1).value or "").strip()
        if value.startswith(label):
            return row
    raise ValueError(f"Could not find section label: {label}")


def source_row_has_data(ws, row):
    return any(ws.cell(row, col).value not in (None, "") for col in (2, 3, 4, 5, 7))


def read_dr_mock_rows(source_path):
    workbook = load_workbook(source_path, data_only=True)
    ws = choose_source_sheet(workbook)
    rows = []

    for row in range(DR_SOURCE_START_ROW, ws.max_row + 1):
        if not source_row_has_data(ws, row):
            continue

        rows.append(
            {
                1: normalize_date(ws.cell(row, 2).value),
                2: normalize_date(ws.cell(row, 3).value),
                4: normalize_date(ws.cell(row, 4).value),
                6: normalize_date(ws.cell(row, 5).value),
                7: normalize_date(ws.cell(row, 7).value),
            }
        )

    return rows


def parse_version_number(text):
    matches = re.findall(r"\d+(?:\.\d+)+", text or "")
    return matches[0] if matches else (text or "").strip()


def parse_cpu(text):
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s+us", text or "")
    return match.group(1) if match else (text or "").strip()


def parse_memory(text):
    total = re.search(r"Total=([^|]+)", text or "")
    used = re.search(r"Used=([^|]+)", text or "")
    return (total.group(1).strip() if total else ""), (used.group(1).strip() if used else "")


def decode_json_list(value):
    if isinstance(value, list):
        return value
    if not value:
        return []
    try:
        return json.loads(value)
    except Exception:
        return [value]


def parse_disk_lines(lines):
    disk_map = {"/": {}, "/app": {}, "/data": {}}

    for line in decode_json_list(lines):
        item = {}
        for part in str(line).split("||"):
            if "=" in part:
                key, value = part.split("=", 1)
                item[key.strip()] = value.strip()

        mount = item.get("Mount")
        if mount in disk_map:
            disk_map[mount] = item

    return disk_map


def parse_process_table(text):
    results = {}

    for line in (text or "").splitlines():
        line = line.strip()
        if not line or line.startswith("NAME"):
            continue

        parts = line.split()
        if not parts:
            continue

        name = parts[0]
        status = " ".join(parts[1:])
        results[name] = status

    return results


def process_result_for(name, mapping):
    process_aliases = {
        "fap-agent": ["fap-agent"],
        "fap-manager": ["fap-manager", "fap-manager-0"],
        "fap-web": ["fap-web", "fap-web2"],
        "fap-haproxy": ["fap-haproxy"],
        "pg": ["pg", "pg-0", "pg-1"],
        "fap-redis": ["fap-redis", "redis", "redis-0"],
        "portainer": ["portainer", "potainer"],
    }

    aliases = process_aliases.get(name, [name])

    for key, status in mapping.items():
        for alias in aliases:
            if key == alias or key.startswith(alias):
                return ("정상" if status.startswith("Up") else "비정상"), status

    return "없음", "docker ps -a 결과 없음"


def parse_csv_lines(text):
    rows = []

    for line in (text or "").splitlines():
        line = line.strip()
        if not line:
            continue

        if line.startswith("조직명,") or line.startswith("조직,"):
            continue

        rows.append([part.strip() for part in line.split(",")])

    return rows


def parse_pipe_lines(text):
    rows = []

    for line in (text or "").splitlines():
        line = line.strip()
        if line:
            rows.append([part.strip() for part in line.split("|")])

    return rows


def normalize_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, dict):
        return ", ".join(f"{k}: {v}" for k, v in value.items())
    return str(value).strip()


def stringify_item(value):
    if isinstance(value, dict):
        return ", ".join(f"{k}: {v}" for k, v in value.items())
    return "" if value is None else str(value)


def set_wrapped_left(cell):
    base = copy(cell.alignment) if cell.alignment else Alignment()
    cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
        text_rotation=base.text_rotation,
        shrink_to_fit=base.shrink_to_fit,
        indent=base.indent,
    )


def set_wrapped_center(cell):
    base = copy(cell.alignment) if cell.alignment else Alignment()
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
        text_rotation=base.text_rotation,
        shrink_to_fit=base.shrink_to_fit,
        indent=base.indent,
    )


def estimate_row_height(*values):
    line_count = 1

    for value in values:
        text = normalize_cell_text(value)
        if not text:
            continue

        wrapped = 0
        for part in text.splitlines() or [""]:
            wrapped += max(1, math.ceil(len(part) / 18))

        line_count = max(line_count, wrapped)

    return max(20, line_count * 18)


def get_servers(payload):
    servers = payload.get("servers") or []

    if isinstance(servers, str):
        try:
            servers = json.loads(servers)
        except Exception:
            servers = []

    if not isinstance(servers, list):
        servers = []

    if not servers:
        servers = [
            {
                "ip": payload.get("first_ip", ""),
                "cpu_output": payload.get("cpu_output", ""),
                "disk_output": payload.get("disk_output", []),
                "memory_output": payload.get("memory_output", ""),
                "process_output": payload.get("process_output", ""),
            }
        ]

        if payload.get("second_ip"):
            servers.append({"ip": payload.get("second_ip", "")})

    return servers


def get_server(payload, index):
    servers = get_servers(payload)

    if index < len(servers) and isinstance(servers[index], dict):
        return servers[index]

    return {}


def fill_top_summary(ws, payload):
    ws.cell(3, 3, payload.get("report_date", ""))
    ws.cell(4, 4, payload.get("first_org", "알티넷솔루션"))
    ws.cell(5, 4, payload.get("second_org", "신세계I&C"))
    ws.cell(4, 6, payload.get("first_name", ""))
    ws.cell(5, 6, payload.get("second_name", ""))
    ws.cell(17, 2, payload.get("special_note", ""))


def fill_host_labels(ws, payload):
    first_ip = str(get_server(payload, 0).get("ip") or payload.get("first_ip", ""))
    second_ip = str(get_server(payload, 1).get("ip") or payload.get("second_ip", ""))

    labels = [
        ("CPU 사용률 확인", "CPU 사용률 확인"),
        ("디스크 사용률 확인", "디스크 사용률 확인"),
        ("메모리 사용률 확인", "메모리 사용률 확인"),
    ]

    for label_text, suffix in labels:
        first_row = find_row_contains(ws, label_text)

        if first_ip:
            ws.cell(first_row, 1, f"    3) {first_ip} {suffix}")

        try:
            second_row = find_row_contains(ws, label_text, start_row=first_row + 1)
            if second_ip:
                ws.cell(second_row, 1, f"    4) {second_ip} {suffix}")
        except ValueError:
            pass

    try:
        process_section = find_section_row(ws, "3. 자동화 서버 관련 프로세스")
        ui_section = find_section_row(ws, "4. 자동화 서버 웹 UI")
        first_process_row = find_row_contains(ws, "1)", start_row=process_section, end_row=ui_section)

        if first_ip:
            ws.cell(first_process_row, 1, f"    1) {first_ip}")

        try:
            second_process_row = find_row_contains(ws, "2)", start_row=first_process_row + 1, end_row=ui_section)
            if second_ip:
                ws.cell(second_process_row, 1, f"    2) {second_ip}")
        except ValueError:
            pass
    except ValueError:
        pass


def fill_version_section(ws, payload):
    version_rows = [
        (30, parse_version_number(payload.get("os_version", "")), payload.get("os_version_eos", ""), payload.get("os_version_note", "")),
        (31, parse_version_number(payload.get("postgres_version", "")), payload.get("db_version_eos", ""), payload.get("db_version_note", "")),
        (32, parse_version_number(payload.get("fap_version", "")), payload.get("fap_version_eos", ""), payload.get("fap_version_note", "")),
        (33, parse_version_number(payload.get("redis_version", "")), payload.get("redis_version_eos", ""), payload.get("redis_version_note", "")),
        (34, parse_version_number(payload.get("docker_version", "")), payload.get("docker_version_eos", ""), payload.get("docker_version_note", "")),
        (35, parse_version_number(payload.get("tomcat_version", "")), payload.get("tomcat_version_eos", ""), payload.get("tomcat_version_note", "")),
        (36, parse_version_number(payload.get("ansible_core_version", "")), payload.get("ansible_core_version_eos", ""), payload.get("ansible_core_version_note", "")),
    ]

    for row, version_value, eos_value, note_value in version_rows:
        ws.cell(row, 2, version_value)
        ws.cell(row, 4, eos_value)
        ws.cell(row, 6, note_value)
        set_wrapped_center(ws.cell(row, 4))
        set_wrapped_left(ws.cell(row, 6))
        ws.row_dimensions[row].height = estimate_row_height(version_value, eos_value, note_value)


def fill_cpu_disk_memory(ws, payload):
    server_rows = [
        {"index": 0, "cpu_row": 48, "disk_rows": [(60, "/"), (61, "/app"), (62, "/data")], "memory_row": 76},
        {"index": 1, "cpu_row": 52, "disk_rows": [(66, "/"), (67, "/app"), (68, "/data")], "memory_row": 80},
    ]

    for mapping in server_rows:
        server = get_server(payload, mapping["index"])
        if not server:
            continue

        cpu = parse_cpu(server.get("cpu_output", ""))
        if cpu:
            ws.cell(mapping["cpu_row"], 1, cpu)
            ws.cell(mapping["cpu_row"], 5, "정상")

        disk_rows = parse_disk_lines(server.get("disk_output", []))
        for rownum, mount in mapping["disk_rows"]:
            item = disk_rows.get(mount, {})
            ws.cell(rownum, 1, mount)

            if item and item.get("Status") != "NOT_FOUND":
                ws.cell(rownum, 2, item.get("Size", ""))
                ws.cell(rownum, 3, item.get("Used", ""))
                ws.cell(rownum, 4, item.get("Use", ""))
                ws.cell(rownum, 5, "정상" if "[FAIL]" not in " ".join(item.values()) else "비정상")
            else:
                ws.cell(rownum, 2, "없음")
                ws.cell(rownum, 3, "없음")
                ws.cell(rownum, 4, "없음")
                ws.cell(rownum, 5, "없음")

        mem_total, mem_used = parse_memory(server.get("memory_output", ""))
        if mem_total or mem_used:
            ws.cell(mapping["memory_row"], 1, mem_total)
            ws.cell(mapping["memory_row"], 2, mem_used)
            ws.cell(mapping["memory_row"], 4, "정상")


def fill_patch_section(ws, payload):
    ws.cell(84, 2, payload.get("os_patch_day", ""))
    ws.cell(84, 4, payload.get("os_vulnerability", ""))
    ws.cell(84, 6, payload.get("os_note", ""))
    ws.cell(85, 2, payload.get("db_patch_day", ""))
    ws.cell(85, 4, payload.get("db_vulnerability", ""))
    ws.cell(85, 6, payload.get("db_note", ""))
    ws.cell(86, 2, payload.get("fap_patch_day", ""))
    ws.cell(86, 4, payload.get("fap_vulnerability", ""))
    ws.cell(86, 6, payload.get("fap_note", ""))


def fill_process_section(ws, payload):
    names = ["fap-agent", "fap-manager", "fap-web", "fap-haproxy", "pg", "fap-redis", "portainer"]

    section_row = find_section_row(ws, "3. 자동화 서버 관련 프로세스")
    next_section_row = find_section_row(ws, "4. 자동화 서버 웹 UI")

    first_label_row = find_row_contains(ws, "1)", start_row=section_row, end_row=next_section_row)
    first_header_row = first_label_row + 1
    first_data_row = first_header_row + 1

    second_label_row = find_row_contains(ws, "2)", start_row=first_data_row, end_row=next_section_row)

    rebuild_dynamic_rows(
        ws=ws,
        start_row=first_data_row,
        next_section_row=second_label_row,
        row_count=len(names),
        template_row=first_data_row,
        max_col=7,
        merge_ranges=[(2, 4), (6, 7)],
    )

    next_section_row = find_section_row(ws, "4. 자동화 서버 웹 UI")
    second_label_row = find_row_contains(ws, "2)", start_row=first_data_row, end_row=next_section_row)
    second_header_row = second_label_row + 1

    normalize_row_merges(ws, second_header_row, merge_ranges=[(2, 4), (6, 7)])
    for col in range(1, 8):
        copy_cell_style(ws, first_header_row, second_header_row, col)

    ws.cell(second_header_row, 1, "번호")
    ws.cell(second_header_row, 2, "프로세스")
    ws.cell(second_header_row, 3, None)
    ws.cell(second_header_row, 4, None)
    ws.cell(second_header_row, 5, "결과")
    ws.cell(second_header_row, 6, "비고")
    ws.cell(second_header_row, 7, None)

    second_data_row = second_header_row + 1
    next_section_row = find_section_row(ws, "4. 자동화 서버 웹 UI")

    rebuild_dynamic_rows(
        ws=ws,
        start_row=second_data_row,
        next_section_row=next_section_row,
        row_count=len(names),
        template_row=second_data_row,
        max_col=7,
        merge_ranges=[(2, 4), (6, 7)],
    )

    for server_index, base_row in enumerate((first_data_row, second_data_row)):
        server = get_server(payload, server_index)
        process_map = parse_process_table(server.get("process_output", "")) if server else {}

        for idx, name in enumerate(names):
            row = base_row + idx
            result, note = process_result_for(name, process_map)

            normalize_row_merges(ws, row, merge_ranges=[(2, 4), (6, 7)])

            get_writable_cell(ws, row, 1).value = idx + 1
            get_writable_cell(ws, row, 2).value = name
            get_writable_cell(ws, row, 5).value = result
            get_writable_cell(ws, row, 6).value = note

            set_wrapped_center(get_writable_cell(ws, row, 1))
            set_wrapped_center(get_writable_cell(ws, row, 2))
            set_wrapped_center(get_writable_cell(ws, row, 5))
            set_wrapped_center(get_writable_cell(ws, row, 6))
            ws.row_dimensions[row].height = estimate_row_height(name, result, note)


def fill_ui_service_section(ws, payload):
    section_row = find_section_row(ws, "4. 자동화 서버 웹 UI")
    next_section_row = find_section_row(ws, "5. 사용자 계정 현황")
    header_row = None

    for row in range(section_row, next_section_row):
        if str(ws.cell(row, 1).value or "").strip() == "번호":
            header_row = row
            break

    if header_row is None:
        return

    for row in range(header_row + 1, next_section_row):
        if ws.cell(row, 1).value not in (None, ""):
            ws.cell(row, 5, "정상")

    ws.cell(header_row + 3, 6, f"현재 등록된 호스트 수 : {payload.get('host_count', '')}")
    ws.cell(header_row + 5, 6, f"현재 등록된 사용자 수 : {payload.get('user_count', '')}")
    ws.cell(header_row + 6, 6, f"현재 등록된 스케줄 수 : {payload.get('schedule_count', '')}")
    ws.cell(header_row + 7, 6, f"현재 등록된 통보 수 : {payload.get('notification_count', '')}")


def fill_user_section(ws, payload):
    user_rows = parse_csv_lines(payload.get("user_account_status", ""))

    section_row = find_section_row(ws, "5. 사용자 계정 현황")
    next_section_row = find_section_row(ws, "6. 자동화 템플릿 점검")
    header_row = find_row_contains(ws, "조직", start_row=section_row, end_row=next_section_row)

    start_row = header_row + 1
    rows_to_fill = user_rows if user_rows else [["", "", "", "", "", ""]]

    rebuild_dynamic_rows(
        ws=ws,
        start_row=start_row,
        next_section_row=next_section_row,
        row_count=len(rows_to_fill),
        template_row=start_row,
        max_col=7,
        merge_ranges=[(6, 7)],
    )

    for idx, row_data in enumerate(rows_to_fill):
        row = start_row + idx

        while len(row_data) < 6:
            row_data.append("")

        ws.cell(row, 1, row_data[0])
        ws.cell(row, 2, row_data[1])
        ws.cell(row, 3, row_data[2])
        ws.cell(row, 4, row_data[3])
        ws.cell(row, 5, row_data[4])
        ws.cell(row, 6, row_data[5])
        ws.cell(row, 7, None)

        set_wrapped_center(ws.cell(row, 1))
        set_wrapped_center(ws.cell(row, 2))
        set_wrapped_center(ws.cell(row, 3))
        set_wrapped_center(ws.cell(row, 4))
        set_wrapped_center(ws.cell(row, 5))
        set_wrapped_center(ws.cell(row, 6))
        ensure_merge(ws, row, 6, row, 7)
        ws.row_dimensions[row].height = estimate_row_height(*row_data[:6])


def fill_template_section(ws, payload):
    title_row = find_section_row(ws, "6. 자동화 템플릿 점검")
    header_row = title_row + 1
    data_row = title_row + 2

    ws.cell(header_row, 1, "전월 수량")
    ws.cell(header_row, 4, "당월 수량")
    normalize_row_merges(ws, header_row, merge_ranges=[(1, 3), (4, 6)])
    normalize_row_merges(ws, data_row, merge_ranges=[(1, 3), (4, 6)])
    clear_row_values(ws, data_row, max_col=7)

    get_writable_cell(ws, header_row, 1).value = "전월 수량"
    get_writable_cell(ws, header_row, 4).value = "당월 수량"
    get_writable_cell(ws, data_row, 1).value = payload.get("previous_template_count", "0")
    get_writable_cell(ws, data_row, 4).value = payload.get("current_template_count", "0")
    set_wrapped_center(get_writable_cell(ws, header_row, 1))
    set_wrapped_center(get_writable_cell(ws, header_row, 4))
    set_wrapped_center(get_writable_cell(ws, data_row, 1))
    set_wrapped_center(get_writable_cell(ws, data_row, 4))


def fill_web_ip_section(ws, payload):
    ip_rows = parse_pipe_lines(payload.get("web_access_ip", ""))

    section_row = find_section_row(ws, "7. 웹 접속 IP 점검")
    next_section_row = find_section_row(ws, "8. DR 워크플로우 점검")
    header_row = find_row_contains(ws, "구분", start_row=section_row, end_row=next_section_row)

    start_row = header_row + 1
    rows_to_fill = ip_rows if ip_rows else [["시스템 접속 허용 IP", "", "", "", ""]]

    normalize_row_merges(ws, header_row, merge_ranges=[(2, 3), (6, 7)])
    rebuild_dynamic_rows(
        ws=ws,
        start_row=start_row,
        next_section_row=next_section_row,
        row_count=len(rows_to_fill),
        template_row=start_row,
        max_col=7,
        merge_ranges=[(1, 4), (5, 7)],
    )

    for idx, row_data in enumerate(rows_to_fill):
        row = start_row + idx

        while len(row_data) < 5:
            row_data.append("")

        left_text = row_data[0]
        if any(row_data[1:4]):
            detail = " / ".join([x for x in row_data[1:4] if x])
            left_text = f"{left_text} | {detail}" if left_text else detail
        right_text = row_data[4]

        ws.cell(row, 1, left_text)
        ws.cell(row, 2, None)
        ws.cell(row, 3, None)
        ws.cell(row, 4, None)
        ws.cell(row, 5, right_text)
        ws.cell(row, 6, None)
        ws.cell(row, 7, None)

        set_wrapped_left(ws.cell(row, 1))
        set_wrapped_left(ws.cell(row, 5))
        ensure_merge(ws, row, 1, row, 4)
        ensure_merge(ws, row, 5, row, 7)
        ws.row_dimensions[row].height = estimate_row_height(*row_data[:5])


def fill_dr_workflow_section(ws, payload):
    title_row = find_section_row(ws, "8. DR 워크플로우 점검")
    header_row = title_row + 1
    mock_row = title_row + 2
    cutover_row = title_row + 3
    manual_row = title_row + 4

    normalize_row_merges(ws, header_row, merge_ranges=[(1, 2), (3, 4), (5, 6)])
    for row in (mock_row, cutover_row, manual_row):
        normalize_row_merges(ws, row, merge_ranges=[(1, 2), (3, 4), (5, 6)])
        clear_row_values(ws, row, max_col=7)

    get_writable_cell(ws, header_row, 1).value = "작업 대상"
    get_writable_cell(ws, header_row, 3).value = "워크플로우 수량(전월)"
    get_writable_cell(ws, header_row, 5).value = "워크플로우 수량(당월)"

    ws.cell(mock_row, 1, "모의훈련")
    ws.cell(cutover_row, 1, "실 전환")
    ws.cell(manual_row, 1, "IP 수동전환")

    ws.cell(mock_row, 3, payload.get("previous_dr_mock_count", "0"))
    ws.cell(mock_row, 5, payload.get("dr_mock_count", "0"))
    ws.cell(cutover_row, 3, payload.get("previous_dr_cutover_count", "0"))
    ws.cell(cutover_row, 5, payload.get("dr_cutover_count", "0"))
    ws.cell(manual_row, 3, payload.get("previous_dr_manual_ip_count", "0"))
    ws.cell(manual_row, 5, payload.get("dr_manual_ip_count", "0"))

    for row in (mock_row, cutover_row, manual_row):
        set_wrapped_center(ws.cell(row, 1))
        set_wrapped_center(ws.cell(row, 3))
        set_wrapped_center(ws.cell(row, 5))
    set_wrapped_center(get_writable_cell(ws, header_row, 1))
    set_wrapped_center(get_writable_cell(ws, header_row, 3))
    set_wrapped_center(get_writable_cell(ws, header_row, 5))


def fill_work_management(ws, payload):
    current = payload.get("current_work_management", []) or []
    upcoming = payload.get("next_work_management", []) or []

    if isinstance(current, str):
        current = json.loads(current) if current else []

    if isinstance(upcoming, str):
        upcoming = json.loads(upcoming) if upcoming else []

    section_row = find_section_row(ws, "9. 작업 관리")
    next_section_row = find_section_row(ws, DR_SECTION_TITLE)
    header_row = find_row_contains(ws, "당월 진행 내역", start_row=section_row, end_row=next_section_row)

    start_row = header_row + 1
    row_count = max(len(current), len(upcoming), 1)

    normalize_row_merges(ws, header_row, merge_ranges=[(1, 4), (5, 7)])
    rebuild_dynamic_rows(
        ws=ws,
        start_row=start_row,
        next_section_row=next_section_row,
        row_count=row_count,
        template_row=start_row,
        max_col=7,
        merge_ranges=[(1, 4), (5, 7)],
    )

    for idx in range(row_count):
        row = start_row + idx
        current_value = stringify_item(current[idx]) if idx < len(current) else ""
        upcoming_value = stringify_item(upcoming[idx]) if idx < len(upcoming) else ""

        ws.cell(row, 1, current_value)
        ws.cell(row, 2, None)
        ws.cell(row, 3, None)
        ws.cell(row, 4, None)
        ws.cell(row, 5, upcoming_value)
        ws.cell(row, 6, None)
        ws.cell(row, 7, None)

        set_wrapped_left(ws.cell(row, 1))
        set_wrapped_left(ws.cell(row, 5))
        ensure_merge(ws, row, 1, row, 4)
        ensure_merge(ws, row, 5, row, 7)
        ws.row_dimensions[row].height = estimate_row_height(current_value, upcoming_value)


def fill_dr_mock_section(ws, rows):
    if not rows:
        return

    title_row = find_row_by_first_column(ws, DR_SECTION_TITLE)
    start_row = title_row + 2
    next_section_row = ws.max_row + 1

    header_row = title_row + 1
    normalize_row_merges(ws, header_row, merge_ranges=[(2, 3), (4, 5)])
    rebuild_dynamic_rows(
        ws=ws,
        start_row=start_row,
        next_section_row=next_section_row,
        row_count=len(rows),
        template_row=start_row,
        max_col=7,
        merge_ranges=[(2, 3), (4, 5)],
    )

    for offset, row_data in enumerate(rows):
        row = start_row + offset

        ws.cell(row, 1, row_data.get(1))
        ws.cell(row, 2, row_data.get(2))
        ws.cell(row, 3, None)
        ws.cell(row, 4, row_data.get(4))
        ws.cell(row, 5, None)
        ws.cell(row, 6, row_data.get(6))
        ws.cell(row, 7, row_data.get(7))

        for col in (1, 2, 4):
            set_wrapped_left(ws.cell(row, col))

        for col in (6, 7):
            set_wrapped_center(ws.cell(row, col))

        ensure_merge(ws, row, 2, row, 3)
        ensure_merge(ws, row, 4, row, 5)
        ws.row_dimensions[row].height = estimate_row_height(
            row_data.get(1),
            row_data.get(2),
            row_data.get(4),
            row_data.get(6),
            row_data.get(7),
        )


def generate_report(template_path, input_payload, output_path):
    workbook = load_workbook(template_path)
    ws = workbook[TARGET_SHEET] if TARGET_SHEET in workbook.sheetnames else workbook.active

    fill_top_summary(ws, input_payload)
    fill_host_labels(ws, input_payload)
    fill_version_section(ws, input_payload)
    fill_cpu_disk_memory(ws, input_payload)
    fill_patch_section(ws, input_payload)
    fill_process_section(ws, input_payload)
    fill_ui_service_section(ws, input_payload)
    fill_user_section(ws, input_payload)
    fill_template_section(ws, input_payload)
    fill_web_ip_section(ws, input_payload)
    fill_dr_workflow_section(ws, input_payload)
    fill_work_management(ws, input_payload)

    dr_mock_source_path = input_payload.get("dr_mock_source_path")
    if dr_mock_source_path and os.path.exists(dr_mock_source_path):
        fill_dr_mock_section(ws, read_dr_mock_rows(dr_mock_source_path))

    fill_host_labels(ws, input_payload)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: generate_server_report.py <template_xlsx> <input_json> <output_xlsx>")

    template_path = sys.argv[1]
    payload = load_input(sys.argv[2])
    output_path = sys.argv[3]

    saved = generate_report(template_path, payload, output_path)
    print(saved)


if __name__ == "__main__":
    main()
