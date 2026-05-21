import json
import os
import sys
from copy import copy
from datetime import datetime
from typing import Any, Dict, List

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import PatternFill
except ImportError as exc:
    print(
        "ERROR: openpyxl 패키지가 필요합니다. "
        "예: pip install openpyxl",
        file=sys.stderr,
    )
    raise

## 프로젝트명 변경시 수정할 곳
TEMPLATE_PATH = "/fap/ansible/projects/[네트워크]_FAP_취약점점검/templates/check_templates_260428.xlsx"
DEFAULT_OUTPUT_DIR = "/fap/report/취약점_점검"

SHEET_NAME = "Sheet1"
HEADER_ROW = 2
DATA_START_ROW = 3
START_COL = 2  # B
END_COL = 5    # E
RESULT_COL = 4  # D
RESULT_COL_WIDTH = 100
MAX_RESULT_TEXT_LENGTH = 5000
ACTION_REQUIRED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FCE4D6",
)


def load_input(arg: str) -> Dict[str, Any]:
    if os.path.isfile(arg):
        with open(arg, "r", encoding="utf-8") as file:
            return json.load(file)
    return json.loads(arg)


def extract_content_value(payload: Dict[str, Any]) -> Dict[str, Any]:
    if "content_value" in payload:
        return payload["content_value"] or {}
    if "job_info" in payload and isinstance(payload["job_info"], dict):
        return payload["job_info"].get("content_value", {}) or {}
    if "data_list" in payload:
        return payload
    return {}


def coerce_result_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        text = "\n".join("" if item is None else str(item) for item in value)
    else:
        text = str(value)

    if len(text) <= MAX_RESULT_TEXT_LENGTH:
        return text

    omitted = len(text) - MAX_RESULT_TEXT_LENGTH
    return f"{text[:MAX_RESULT_TEXT_LENGTH]}\n... (truncated, {omitted} chars omitted)"


def get_result_text(entry: Dict[str, Any]) -> str:
    for key in ("result", "output", "stdout", "message", "stderr"):
        if key in entry and entry.get(key) is not None:
            return coerce_result_text(entry.get(key)).strip()
    return ""


def sanitize_for_excel(value: str) -> str:
    if value and value[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def infer_action_required(entry: Dict[str, Any], result_text: str) -> str:
    explicit = entry.get("action_required")
    if explicit is None:
        explicit = entry.get("need_action")
    if explicit is None:
        explicit = entry.get("is_action_required")

    if explicit is not None:
        if isinstance(explicit, bool):
            return "필요" if explicit else "불필요"
        text = str(explicit).strip().lower()
        if text in {"y", "yes", "true", "1", "need", "required", "필요"}:
            return "필요"
        if text in {"n", "no", "false", "0", "none", "불필요"}:
            return "불필요"

    upper_text = result_text.upper()
    if "[FAIL]" in upper_text or "FAIL" in upper_text:
        return "필요"
    if "[PASS]" in upper_text and "[FAIL]" not in upper_text:
        return "불필요"

    rc = entry.get("rc")
    if isinstance(rc, int):
        return "필요" if rc != 0 else "불필요"

    return ""


def normalize_entries(content_value: Dict[str, Any]) -> List[Dict[str, Any]]:
    data_list = content_value.get("data_list", [])
    normalized: List[Dict[str, Any]] = []

    for index, entry in enumerate(data_list, start=1):
        if not isinstance(entry, dict):
            continue

        script_name = (
            entry.get("script_name")
            or entry.get("cmd_name")
            or entry.get("command")
            or entry.get("cmd")
            or f"script_{index}"
        )
        result_text = get_result_text(entry)
        sequence = (
            entry.get("seq")
            or entry.get("cmd_id")
            or entry.get("number")
            or index
        )

        normalized.append(
            {
                "number": sequence,
                "script_name": str(script_name),
                "result_text": result_text,
                "action_required": infer_action_required(entry, result_text),
            }
        )

    return normalized


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(START_COL, END_COL + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)

        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_existing_rows(ws, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(START_COL, END_COL + 1):
            ws.cell(row=row, column=col).value = None


def ensure_capacity(ws, needed_rows: int) -> int:
    current_capacity = max(ws.max_row - DATA_START_ROW + 1, 0)
    last_row = max(ws.max_row, DATA_START_ROW)

    if needed_rows <= current_capacity:
        return max(DATA_START_ROW + needed_rows - 1, DATA_START_ROW)

    template_row = DATA_START_ROW
    for _ in range(needed_rows - current_capacity):
        last_row += 1
        copy_row_style(ws, template_row, last_row)

    return last_row


def estimate_wrapped_line_count(text: str, width_chars: int) -> int:
    if not text:
        return 1

    normalized_width = max(width_chars, 1)
    total = 0
    for line in str(text).splitlines() or [""]:
        line_length = max(len(line), 1)
        total += (line_length - 1) // normalized_width + 1
    return max(total, 1)


def adjust_layout(ws, rows: List[Dict[str, Any]]) -> None:
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 6)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 22)
    ws.column_dimensions["D"].width = RESULT_COL_WIDTH
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width or 0, 14)

    usable_width = RESULT_COL_WIDTH - 4
    for offset, row_data in enumerate(rows):
        row_idx = DATA_START_ROW + offset
        result_lines = estimate_wrapped_line_count(
            row_data["result_text"],
            usable_width,
        )
        script_lines = estimate_wrapped_line_count(row_data["script_name"], 18)
        row_lines = max(result_lines, script_lines, 1)
        ws.row_dimensions[row_idx].height = min(max(24, row_lines * 18), 409)


def write_rows(ws, rows: List[Dict[str, Any]]) -> None:
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    centered_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    end_row = ensure_capacity(ws, len(rows))
    clear_existing_rows(ws, DATA_START_ROW, end_row)

    for offset, row_data in enumerate(rows):
        row_idx = DATA_START_ROW + offset
        ws.cell(row=row_idx, column=2, value=row_data["number"])
        ws.cell(row=row_idx, column=3, value=row_data["script_name"])
        ws.cell(
            row=row_idx,
            column=4,
            value=sanitize_for_excel(row_data["result_text"]),
        )
        action_cell = ws.cell(row=row_idx, column=5, value=row_data["action_required"])
        if row_data["action_required"] == "필요":
            action_cell.fill = ACTION_REQUIRED_FILL

        for col in range(START_COL, END_COL + 1):
            ws.cell(row=row_idx, column=col).alignment = wrap_alignment

        ws.cell(row=row_idx, column=2).alignment = centered_alignment
        ws.cell(row=row_idx, column=3).alignment = centered_alignment
        ws.cell(row=row_idx, column=5).alignment = centered_alignment

    adjust_layout(ws, rows)


def generate_output_path(output_path: str = "") -> str:
    if output_path:
        return output_path

    os.makedirs(DEFAULT_OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(DEFAULT_OUTPUT_DIR, f"취약점_점검_보고서_{timestamp}.xlsx")


def cleanup_existing_reports(output_path: str) -> None:
    if not output_path:
        return

    report_dir = os.path.dirname(output_path)
    report_name = os.path.basename(output_path)
    report_prefix = "취약점_점검_보고서_"

    if not os.path.isdir(report_dir):
        return

    for name in os.listdir(report_dir):
        if name == report_name:
            continue
        if not name.startswith(report_prefix) or not name.endswith(".xlsx"):
            continue

        old_report_path = os.path.join(report_dir, name)
        if os.path.isfile(old_report_path):
            os.remove(old_report_path)


def generate_report(payload: Dict[str, Any], output_path: str = "") -> Dict[str, Any]:
    content_value = extract_content_value(payload)
    rows = normalize_entries(content_value)
    if not rows:
        return {
            "is_success": False,
            "file_path": None,
            "message": "data_list에서 보고서로 쓸 데이터를 찾지 못했습니다.",
        }

    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

    write_rows(worksheet, rows)

    final_output_path = generate_output_path(output_path)
    os.makedirs(os.path.dirname(final_output_path), exist_ok=True)
    cleanup_existing_reports(final_output_path)
    workbook.save(final_output_path)

    return {
        "is_success": True,
        "file_path": final_output_path,
        "row_count": len(rows),
        "message": "작업 보고서 생성 완료",
    }


def main() -> int:
    if len(sys.argv) < 2:
        print(
            "ERROR: JSON 문자열 또는 JSON 파일 경로를 인자로 넘겨주세요.",
            file=sys.stderr,
        )
        return 1

    try:
        payload = load_input(sys.argv[1])
        output_path = sys.argv[2] if len(sys.argv) > 2 else ""
        result = generate_report(payload, output_path)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["is_success"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
