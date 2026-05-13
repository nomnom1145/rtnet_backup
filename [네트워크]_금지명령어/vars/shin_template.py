import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import os
import json
import sys
from datetime import datetime

TEMPLATE_PATH = '/fap/report/make_report/templates/report_template_SH_network.xlsx'

def get_column_letter(col_idx):
    if col_idx < 1:
        raise ValueError("Column index must be 1 or greater")
    letters = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

def generate_report(content_value, report_path):
    try:
        wb = load_workbook(TEMPLATE_PATH, data_only=True)
    except Exception as e:
        print(f"ERROR: Failed to load template {TEMPLATE_PATH}: {e}", file=sys.stderr)
        return {"is_success": False, "file_path": None}

    sheet = wb.active

    header_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    content_aligned = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    data_list = content_value.get('data_list', [])
    if not data_list:
        print("ERROR: No data to write in report", file=sys.stderr)
        return {"is_success": False, "file_path": None}

    hosts = {}
    for entry in data_list:
        host = entry.get('host')
        cmd = entry.get('command', '')
        result = (entry.get('result') or '').strip()

        # Skip N/A or empty results
        if not result or result == 'N/A':
            continue

        # Extract host_name from invocation.module_args.fap_vars.host_name if available
        host_name = entry.get('invocation', {}).get('module_args', {}).get('fap_vars', {}).get('host_name')
        if not host_name:
            host_name = entry.get('fap_vars', {}).get('host_name')
        if not host_name:
            host_name = entry.get('host_name', '')

        if not host:
            continue

        if host not in hosts:
            hosts[host] = {'host_name': host_name, 'results': {}}
        hosts[host]['results'][cmd] = result

    commands = []
    for host_data in hosts.values():
        for cmd in host_data['results'].keys():
            if cmd not in commands:
                commands.append(cmd)

    for col_idx, command in enumerate(commands, start=3):
        col_letter = get_column_letter(col_idx)
        cell = sheet[f'{col_letter}1']
        cell.value = f"{command} 명령어 결과"
        cell.alignment = header_aligned
        cell.border = thin_border

    for row_idx, (host, info) in enumerate(hosts.items(), start=2):
        cell_host = sheet[f'A{row_idx}']
        cell_host.value = host
        cell_host.alignment = content_aligned
        cell_host.border = thin_border

        cell_name = sheet[f'B{row_idx}']
        cell_name.value = info['host_name']
        cell_name.alignment = content_aligned
        cell_name.border = thin_border

        for cmd_idx, cmd in enumerate(commands, start=3):
            col_letter = get_column_letter(cmd_idx)
            cell = sheet[f'{col_letter}{row_idx}']
            val = info['results'].get(cmd, '')
            cell.value = val
            cell.alignment = content_aligned
            cell.border = thin_border

    max_lengths = {get_column_letter(i): 0 for i in range(1, len(commands) + 3)}
    for r in range(1, sheet.max_row + 1):
        for c in range(1, len(commands) + 3):
            col_letter = get_column_letter(c)
            cell = sheet[f'{col_letter}{r}']
            if cell.value:
                length = max(len(str(line)) for line in str(cell.value).split('\n'))
                max_lengths[col_letter] = max(max_lengths[col_letter], length)
    for col_letter, width in max_lengths.items():
        sheet.column_dimensions[col_letter].width = min(width + 2, 50)

    for r in range(2, sheet.max_row + 1):
        max_lines = 1
        for c in range(1, len(commands) + 3):
            col_letter = get_column_letter(c)
            cell = sheet[f'{col_letter}{r}']
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                max_lines = max(max_lines, lines)
        sheet.row_dimensions[r].height = min(max_lines, 20) * 15

    try:
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        wb.save(report_path)
    except Exception as e:
        print(f"ERROR: Failed to save report {report_path}: {e}", file=sys.stderr)
        return {"is_success": False, "file_path": None}

    return {"is_success": True, "file_path": report_path}

def main(job_info):
    content_value = job_info.get("content_value", {})
    job_id = job_info.get("job_id", "unknown")
    today_date = datetime.now().strftime("%Y%m%d")
    report_path = os.path.join("/fap/report/Network_점검", today_date, job_id, 
                                f"Network_점검_{today_date}_{job_id}.xlsx")  # 보고서 위치의 경우 변경 가능
    result = generate_report(content_value, report_path)
    sys.exit(0 if result.get("is_success") else 1)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        try:
            if os.path.isfile(arg):
                with open(arg, 'r') as f:
                    job_info = json.load(f)
            else:
                job_info = json.loads(arg)
        except Exception as e:
            print(f"ERROR: Invalid JSON input: {e}", file=sys.stderr)
            sys.exit(1)
        main(job_info)
    else:
        print("ERROR: Missing job_info argument", file=sys.stderr)
        sys.exit(1)
