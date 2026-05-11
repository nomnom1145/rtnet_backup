import json
import os
import re
from datetime import datetime

from ansible.plugins.filter.core import FilterModule as CoreFilterModule
from ansible.utils.display import Display
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


display = Display()

DEFAULT_REPORT_TYPE = "SHIN_server_RedirFS"
HEADERS = ["Hostname", "IP", "Command", "Result", "Failure Reason"]


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(text(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, indent=2)
    return str(value)


def first_value(*values) -> str:
    for value in values:
        normalized = text(value).strip()
        if normalized:
            return normalized
    return ""


def normalize_rows(json_data):
    """
    json_data는 아래 둘 다 허용:
    1) list
       - merged 같은 점검 결과 리스트
    2) dict
       - {"content_value": {"data_list": [...]}} 형태
    """

    if isinstance(json_data, list):
        data_list = json_data
    elif isinstance(json_data, dict):
        data_list = json_data.get("content_value", {}).get("data_list", [])
    else:
        data_list = []

    rows = []

    for entry in data_list:
        if not isinstance(entry, dict):
            continue

        rows.append(
            {
                "hostname": first_value(
                    entry.get("hostname"),
                    entry.get("host_name"),
                    entry.get("real_host_name"),
                ),
                "ip": first_value(
                    entry.get("ip"),
                    entry.get("host"),
                ),
                "command": first_value(
                    entry.get("command"),
                    entry.get("cmd"),
                    entry.get("cmd_name"),
                ),
                "result": first_value(
                    entry.get("version"),
                    entry.get("result"),
                    entry.get("output"),
                ),
                "failure_reason": first_value(
                    entry.get("failure_reason"),
                    entry.get("message"),
                ),
            }
        )

    rows.sort(key=lambda item: (item["hostname"], item["ip"], item["command"]))
    return rows


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned or DEFAULT_REPORT_TYPE


def build_output_path(json_data, file_path: str) -> str:
    """
    UI/Playbook에서 file_path를 넘기면 그 경로를 최우선으로 사용.
    - file_path가 .xlsx면 그대로 파일 경로로 사용
    - file_path가 디렉토리면 그 아래에 파일명 자동 생성
    """

    if not file_path:
        raise ValueError("file_path is required")

    file_path = text(file_path).strip()

    if file_path.lower().endswith(".xlsx"):
        output_path = file_path
    else:
        today = datetime.now().strftime("%Y%m%d")

        report_type = DEFAULT_REPORT_TYPE
        if isinstance(json_data, dict):
            report_type = first_value(json_data.get("report_type")) or DEFAULT_REPORT_TYPE
        elif isinstance(json_data, list) and json_data:
            report_type = first_value(json_data[0].get("report_type")) or DEFAULT_REPORT_TYPE

        report_type = safe_filename(report_type)
        output_path = os.path.join(file_path, f"{report_type}_{today}.xlsx")

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    return output_path


def autosize(ws):
    widths = {}

    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            max_line = max((len(line) for line in value.splitlines()), default=0)
            widths[cell.column] = max(widths.get(cell.column, 0), max_line)

    for column_index, width in widths.items():
        ws.column_dimensions[
            ws.cell(row=1, column=column_index).column_letter
        ].width = min(width + 4, 80)


def write_workbook(rows, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RedirFS"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    fail_fill = PatternFill(fill_type="solid", fgColor="F8CBAD")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["hostname"],
            row["ip"],
            row["command"],
            row["result"],
            row["failure_reason"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.alignment = body_alignment
            cell.border = border

        ws.row_dimensions[row_index].height = 22

    ws.freeze_panes = "A2"
    autosize(ws)
    wb.save(output_path)


def create_version_report(json_data, file_path: str) -> bool:
    try:
        rows = normalize_rows(json_data)

        if not rows:
            display.display("ERROR: No rows to write")
            return False

        output_path = build_output_path(json_data, file_path)

        write_workbook(rows, output_path)

        display.display(f"RedirFS 보고서가 생성되었습니다: {output_path}")
        return True

    except Exception as e:
        import traceback

        display.display(f"RedirFS 보고서 생성 중 오류 발생: {e}")
        traceback.print_exc()
        return False



class FilterModule(CoreFilterModule):
    def filters(self):
        return {
            "make_version_report": self.execute_filter,
            "make_redirfs_report": self.execute_filter,
        }

    def execute_filter(self, json_data, file_path):
        return create_version_report(json_data, file_path)